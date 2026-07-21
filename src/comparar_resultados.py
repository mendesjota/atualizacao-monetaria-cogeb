"""
Compara o Excel gerado pelo pipeline com os valores de referência do ODS.
"""
from __future__ import annotations

import json
import os
import sys
from pathlib import Path

os.environ["PYTHONIOENCODING"] = "utf-8"
sys.stdout.reconfigure(encoding="utf-8", errors="replace") if hasattr(sys.stdout, "reconfigure") else None

from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parent.parent
REF_PATH = RAIZ / "dados" / "validacao" / "referencia_ods.json"


def moeda_para_float(txt):
    if txt is None:
        return 0.0
    if isinstance(txt, (int, float)):
        return float(txt)
    return float(str(txt).replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".").strip() or "0")


def load_reference():
    with open(REF_PATH, encoding="utf-8") as f:
        return json.load(f)


def extrair_parcelas_xlsx(caminho: Path | str) -> dict:
    """Extrai parcelas do Excel gerado, agrupadas por beneficiário."""
    caminho = Path(caminho)
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active

    resultado: dict[str, list[dict]] = {}
    current_key = None
    current_data = []
    current_nome = None
    current_mat = None

    for row in ws.iter_rows(min_row=1, values_only=True):
        vals = list(row)

        # Detect NOME: line
        if vals[0] and "NOME:" in str(vals[0]).upper():
            if current_key and current_data:
                resultado[current_key] = current_data

            nome_full = str(vals[1] or "")
            current_mat = ""
            import re
            m = re.search(r"MATR[^A-Za-z0-9]*CULA\s+(\S+)", nome_full, re.IGNORECASE)
            if m:
                current_mat = m.group(1)
            current_nome = re.sub(r"\s*[-–—]\s*MATR[^A-Za-z0-9]*CULA\s+\S+", "", nome_full, flags=re.IGNORECASE).strip()
            current_key = f"{current_nome}||{current_mat}"
            current_data = []
            continue

        # Data line: check if first cell contains a date
        from datetime import datetime, date as date_type
        if vals[0] and isinstance(vals[0], (datetime, date_type)):
            d = vals[0]
            comp_str = f"{d.year:04d}-{d.month:02d}"
            valor_original = float(vals[1] or 0)
            desc = str(vals[2] or "")
            valor_corrigido = float(vals[6] or 0) if len(vals) > 6 else 0
            if "TOTAL" not in desc.upper() and "Total" not in str(vals[0] or ""):
                current_data.append({
                    "competencia": comp_str,
                    "valor_original": valor_original,
                    "valor_corrigido": valor_corrigido,
                })

    if current_key and current_data:
        resultado[current_key] = current_data

    wb.close()
    return resultado


def comparar(resultado: dict, referencia: dict, tolerancia: float = 0.02):
    """Compara parcelas extraídas com referência. tolerancia em reais."""
    total_parcelas = 0
    total_ok = 0
    total_diff = 0
    divergencias = []

    for key, ref_parcelas in referencia.items():
        if key not in resultado:
            continue
        res_parcelas = resultado[key]
        total_parcelas += len(ref_parcelas)

        # Index by competencia (lista para suportar múltiplas parcelas no mesmo mês)
        ref_by_comp: dict[str, list[dict]] = {}
        res_by_comp: dict[str, list[dict]] = {}
        for p in ref_parcelas:
            ref_by_comp.setdefault(p["competencia"], []).append(p)
        for p in res_parcelas:
            res_by_comp.setdefault(p["competencia"], []).append(p)

        for comp, ref_list in ref_by_comp.items():
            res_list = res_by_comp.get(comp, [])
            for i, ref_p in enumerate(ref_list):
                if i >= len(res_list):
                    divergencias.append(f"{key} | {comp} | Parcela {i+1} não encontrada no resultado")
                    total_diff += 1
                    continue

                res_p = res_list[i]
                diff_original = abs(ref_p["valor_original"] - res_p["valor_original"])
                diff_corrigido = abs(ref_p["valor_corrigido"] - res_p["valor_corrigido"])

                if diff_original > tolerancia:
                    divergencias.append(
                        f"{key} | {comp} #{i+1} | Valor Original: esperado={ref_p['valor_original']:.2f}, "
                        f"obtido={res_p['valor_original']:.2f}, diff={diff_original:.2f}"
                    )
                    total_diff += 1
                elif diff_corrigido > tolerancia:
                    divergencias.append(
                        f"{key} | {comp} #{i+1} | Valor Corrigido: esperado={ref_p['valor_corrigido']:.2f}, "
                        f"obtido={res_p['valor_corrigido']:.2f}, diff={diff_corrigido:.2f}"
                    )
                    total_diff += 1
                else:
                    total_ok += 1

            # Parcelas extras no resultado (não esperadas)
            if len(res_list) > len(ref_list):
                for j in range(len(ref_list), len(res_list)):
                    divergencias.append(f"{key} | {comp} | Parcela extra #{j+1} no resultado")
                    total_diff += 1

    return {
        "total_parcelas": total_parcelas,
        "total_ok": total_ok,
        "total_diff": total_diff,
        "divergencias": divergencias,
    }


def main(xlsx_path: str):
    xlsx = Path(xlsx_path)
    if not xlsx.exists():
        print(f"✖ Arquivo não encontrado: {xlsx}")
        return 1

    print(f"→ Lendo referência: {REF_PATH}")
    ref = load_reference()
    print(f"  {len(ref)} beneficiários na referência.")

    print(f"→ Lendo resultado: {xlsx}")
    resultado = extrair_parcelas_xlsx(xlsx)
    print(f"  {len(resultado)} beneficiários encontrados no resultado.")

    print("\n→ Comparando...")
    stats = comparar(resultado, ref)

    print(f"\n  Total parcelas: {stats['total_parcelas']}")
    print(f"  OK:            {stats['total_ok']}")
    print(f"  Divergências:  {stats['total_diff']}")

    if stats["divergencias"]:
        print(f"\n--- DIVERGÊNCIAS ({len(stats['divergencias'])}) ---")
        for d in stats["divergencias"][:50]:
            print(f"  {d}")

    if stats["total_diff"] == 0:
        print("\n✔ Todas as parcelas conferem!")
    else:
        print(f"\n⚠  {stats['total_diff']} parcela(s) com divergência.")

    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python src/comparar_resultados.py <arquivo.xlsx>")
        sys.exit(1)
    sys.exit(main(sys.argv[1]))
