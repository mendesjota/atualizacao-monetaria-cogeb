"""
Leitura da planilha de entrada (.csv ou .xlsx) e escrita do Excel de saída
no formato COGEB — esteticamente idêntico ao template ODS (JAN_-_2026).
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from calculo import Beneficiario, ResumoBeneficiario, parse_competencia

COLUNAS_OBRIGATORIAS = [
    "nome", "matricula", "orgao", "valor_mensal",
    "competencia_inicial", "competencia_final", "data_alvo",
]

# --------------------------------------------------------------------------
# Leitura
# --------------------------------------------------------------------------
def _valor_para_float(texto) -> float:
    if isinstance(texto, (int, float)):
        return float(texto)
    t = str(texto).strip().replace("R$", "").strip()
    if "," in t:
        t = t.replace(".", "").replace(",", ".")
    return float(t or 0)


def _linha_para_beneficiario(linha: dict, num: int) -> Beneficiario:
    faltando = [c for c in COLUNAS_OBRIGATORIAS if not str(linha.get(c, "")).strip()]
    if faltando:
        raise ValueError(f"Linha {num}: faltam colunas {faltando}.")
    return Beneficiario(
        nome=str(linha["nome"]).strip(),
        matricula=str(linha["matricula"]).strip(),
        orgao=str(linha["orgao"]).strip(),
        valor_mensal=_valor_para_float(linha["valor_mensal"]),
        competencia_inicial=parse_competencia(linha["competencia_inicial"]),
        competencia_final=parse_competencia(linha["competencia_final"]),
        data_alvo=parse_competencia(linha["data_alvo"]),
        observacao=str(linha.get("observacao", "") or "").strip(),
    )


def ler_entrada(caminho: str | Path) -> list[Beneficiario]:
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo de entrada não encontrado: {caminho}")
    if caminho.suffix.lower() == ".csv":
        linhas = _ler_csv(caminho)
    elif caminho.suffix.lower() in (".xlsx", ".xlsm"):
        linhas = _ler_xlsx(caminho)
    else:
        raise ValueError(f"Formato não suportado: {caminho.suffix} (use .csv ou .xlsx)")
    beneficiarios = []
    for i, linha in enumerate(linhas, start=2):
        if not any(str(v).strip() for v in linha.values()):
            continue
        beneficiarios.append(_linha_para_beneficiario(linha, i))
    if not beneficiarios:
        raise ValueError("Nenhum beneficiário válido encontrado na planilha.")
    return beneficiarios


def _ler_csv(caminho: Path) -> list[dict]:
    with caminho.open("r", encoding="utf-8-sig", newline="") as f:
        amostra = f.read(2048)
        f.seek(0)
        sep = ";" if amostra.count(";") >= amostra.count(",") else ","
        leitor = csv.DictReader(f, delimiter=sep)
        leitor.fieldnames
        return [{(k or "").strip(): v for k, v in linha.items()} for linha in leitor]


def _ler_xlsx(caminho: Path) -> list[dict]:
    wb = load_workbook(caminho, data_only=True, read_only=True)
    ws = wb.active
    linhas_iter = ws.iter_rows(values_only=True)
    cabecalho = [str(c).strip() if c is not None else "" for c in next(linhas_iter)]
    registros = [dict(zip(cabecalho, valores)) for valores in linhas_iter]
    wb.close()
    return registros


# --------------------------------------------------------------------------
# Constantes de estilo — ODS JAN_-_2026
# --------------------------------------------------------------------------
_MOEDA = 'R$ #,##0.00'
_FATOR = '0.000000'
_DATA_BR = 'DD/MM/YYYY'

_BRANCO = "FFFFFFFF"
_PRETO = "FF000000"
_VERMELHO = "FFFF0000"
_VERDE = "FF00FF00"
_AZUL_13 = "FFB4C7DC"

_BORDA_FINA = Border(
    left=Side(style="thin", color=_PRETO),
    right=Side(style="thin", color=_PRETO),
    top=Side(style="thin", color=_PRETO),
    bottom=Side(style="thin", color=_PRETO),
)
_BORDA_DADOS = Border(
    left=Side(style="thin", color=_PRETO),
    right=Side(style="thin", color=_PRETO),
    top=Side(style="thin", color=_PRETO),
    bottom=Side(style="thin", color=_PRETO),
)


_FONTE_NOME_LABEL = Font(name="Calibri", size=14, bold=True)
_FONTE_NOME_VALOR = Font(name="Calibri", size=14, bold=True, color=_VERMELHO)
_FONTE_ORGAO_LABEL = Font(name="Calibri", size=16, bold=True)
_FONTE_ORGAO_VALOR = Font(name="Calibri", size=16, bold=True, color=_VERMELHO)
_FONTE_CAB = Font(name="Calibri", bold=True)
_FONTE_DADOS = Font(name="Liberation Serif", size=10)
_FONTE_TOTAL = Font(name="Calibri", bold=True)
_FONTE_RUB_TITULO = Font(name="Calibri", bold=True)
_FONTE_RUB_VALOR = Font(name="Calibri", size=12, bold=True)

_ALINH_CENTRO = Alignment(horizontal="center", vertical="center")
_ALINH_DIR = Alignment(horizontal="right", vertical="center")
_ALINH_ESQ = Alignment(horizontal="left", vertical="center")
_ALINH_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALINH_CAB = Alignment(horizontal="center", vertical="center", wrap_text=True)

_FILL_VERDE = PatternFill("solid", fgColor=_VERDE)
_FILL_AZUL_13 = PatternFill("solid", fgColor=_AZUL_13)
_FILL_BRANCO = PatternFill("solid", fgColor=_BRANCO)

_MESES_ABR = {1: "jan", 2: "fev", 3: "mar", 4: "abr", 5: "mai", 6: "jun",
              7: "jul", 8: "ago", 9: "set", 10: "out", 11: "nov", 12: "dez"}


def _mes_ano_abrev(d) -> str:
    return f"{_MESES_ABR[d.month]}/{d.year % 100:02d}"


# --------------------------------------------------------------------------
# Escrita
# --------------------------------------------------------------------------
def escrever_saida(
    resumos: list[ResumoBeneficiario],
    caminho: str | Path,
    indice: str = "INPC",
) -> Path:
    caminho = Path(caminho)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    linha = 1
    for r in resumos:
        _bloco_beneficiario(ws, linha, r)
        linha += len(r.parcelas) + 8
    _config_colunas(ws)
    _config_impressao(ws)
    wb.save(caminho)
    return caminho


def _config_colunas(ws) -> None:
    widths = {"A": 11.5, "B": 12.3, "C": 22.0, "D": 11.5,
              "E": 9.0, "F": 10.0, "G": 9.0,
              "H": 5.0, "I": 5.0, "J": 5.0}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _config_impressao(ws) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _escrever_celula(ws, linha, col, valor=None, font=None, number_format=None,
                     alignment=None, fill=None, border=None):
    c = ws.cell(row=linha, column=col, value=valor)
    if font: c.font = font
    if number_format: c.number_format = number_format
    if alignment: c.alignment = alignment
    if fill: c.fill = fill
    if border: c.border = border
    return c


def _bloco_beneficiario(ws, linha, r) -> None:
    # ── NOME / ÓRGÃO ──
    _escrever_celula(ws, linha, 1, "NOME:", _FONTE_NOME_LABEL,
                     alignment=_ALINH_ESQ, fill=_FILL_BRANCO, border=_BORDA_FINA)
    _escrever_celula(ws, linha, 2, f"{r.nome} - MATRÍCULA {r.matricula}",
                     _FONTE_NOME_VALOR, alignment=_ALINH_CENTRO, border=_BORDA_FINA)
    _escrever_celula(ws, linha, 3, None, border=_BORDA_FINA)
    _escrever_celula(ws, linha, 4, None, border=_BORDA_FINA)
    _escrever_celula(ws, linha, 5, None, border=_BORDA_FINA)
    ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=5)
    _escrever_celula(ws, linha, 6, "ÓRGÃO:", _FONTE_ORGAO_LABEL,
                     alignment=_ALINH_CENTRO, fill=_FILL_BRANCO, border=_BORDA_FINA)
    _escrever_celula(ws, linha, 7, r.orgao, _FONTE_ORGAO_VALOR,
                     alignment=_ALINH_CENTRO, fill=_FILL_BRANCO, border=_BORDA_FINA)
    linha += 1

    # ── Cabeçalho da tabela ──
    titulos = [
        "Data do Valor Original", "Valor Original", "Descrição",
        "Data de Atualização", "Atualização Monetária",
        "Fator de Atualização", "Valor Corrigido",
    ]
    for col, t in enumerate(titulos, start=1):
        _escrever_celula(ws, linha, col, t, _FONTE_CAB,
                         alignment=_ALINH_CAB, border=_BORDA_FINA)
    linha += 1

    # ── Todos os dados na ordem de _ordenar_ods ──
    linha_ini = linha
    for p in r.parcelas:
        _escrever_linha_dados(ws, linha, p, r, tipo=p.tipo)
        linha += 1
    linha_fim_tudo = linha - 1

    # ── Total ──
    primeiro = linha_ini
    ultimo = linha_fim_tudo
    tem_dados = ultimo >= primeiro

    _escrever_celula(ws, linha, 1, "Total:", _FONTE_TOTAL,
                     alignment=_ALINH_CENTRO, fill=_FILL_BRANCO, border=_BORDA_FINA)
    if tem_dados:
        _escrever_celula(ws, linha, 2, f"=SUM(B{primeiro}:B{ultimo})",
                         _FONTE_TOTAL, _MOEDA, _ALINH_DIR, border=_BORDA_FINA)
        _escrever_celula(ws, linha, 5, f"=SUM(E{primeiro}:E{ultimo})",
                         _FONTE_TOTAL, _MOEDA, _ALINH_DIR, border=_BORDA_FINA)
        _escrever_celula(ws, linha, 7, f"=SUM(G{primeiro}:G{ultimo})",
                         _FONTE_TOTAL, _MOEDA, _ALINH_DIR, border=_BORDA_FINA)
    linha += 1

    # ── RUBRICA ──
    mes_alvo = _mes_ano_abrev(r.data_alvo)

    _escrever_celula(ws, linha, 1, "RUBRICA", _FONTE_RUB_TITULO,
                     alignment=_ALINH_CENTRO, fill=_FILL_VERDE, border=_BORDA_FINA)
    _escrever_celula(ws, linha, 2, "VALOR", _FONTE_RUB_TITULO,
                     alignment=_ALINH_CENTRO, fill=_FILL_VERDE, border=_BORDA_FINA)
    _escrever_celula(ws, linha, 3, "Mês/Ano da Devolução", _FONTE_RUB_TITULO,
                     alignment=_ALINH_CENTRO, border=_BORDA_FINA)
    _escrever_celula(ws, linha, 4, "Data do Requerimento", _FONTE_RUB_TITULO,
                     alignment=_ALINH_CENTRO, border=_BORDA_FINA)
    linha += 1

    # Pré-calcula valores das rubricas a partir das parcelas
    sum_regular = round(sum(p.valor_corrigido for p in r.parcelas if p.tipo != "decimo_terceiro"), 2)
    sum_decimo = round(sum(p.valor_corrigido for p in r.parcelas if p.tipo == "decimo_terceiro"), 2)
    sum_correcao = round(sum(p.correcao for p in r.parcelas), 2)

    # 30920 — regular + diferença (exclui 13º)
    _escrever_celula(ws, linha, 1, "30920 - SEGURIDADE SOCIAL", _FONTE_RUB_TITULO,
                     alignment=_ALINH_CENTRO, fill=_FILL_VERDE, border=_BORDA_FINA)
    _escrever_celula(ws, linha, 2, sum_regular,
                     _FONTE_RUB_VALOR, _MOEDA, _ALINH_CENTRO, _FILL_VERDE, _BORDA_FINA)
    _escrever_celula(ws, linha, 3, mes_alvo, _FONTE_RUB_TITULO,
                     alignment=_ALINH_CENTRO, border=_BORDA_FINA)
    linha += 1

    # 30923 — 13º salário
    _escrever_celula(ws, linha, 1, "30923 - SEG. SOC. - 13º SAL.", _FONTE_RUB_TITULO,
                     alignment=_ALINH_CENTRO, fill=_FILL_VERDE, border=_BORDA_FINA)
    _escrever_celula(ws, linha, 2, sum_decimo,
                     _FONTE_RUB_VALOR, _MOEDA, _ALINH_CENTRO, _FILL_VERDE, _BORDA_FINA)
    linha += 1

    # 20827 — correção monetária total
    _escrever_celula(ws, linha, 1, "20827 - ATUAL. MONETÁRIA.", _FONTE_RUB_TITULO,
                     alignment=_ALINH_CENTRO, fill=_FILL_VERDE, border=_BORDA_FINA)
    _escrever_celula(ws, linha, 2, sum_correcao,
                     _FONTE_RUB_VALOR, _MOEDA, _ALINH_CENTRO, _FILL_VERDE, _BORDA_FINA)
    linha += 1

    # ── Blank entre blocos ──
    linha += 1



def _escrever_linha_dados(ws, linha, p, r, tipo: str) -> None:
    if tipo == "decimo_terceiro":
        desc = f"DEVOLUÇÃO DA SEGURIDADE SOCIAL 13º SAL - {p.nome} - MATRÍCULA {r.matricula}"
        fill = _FILL_AZUL_13
    elif tipo == "diferenca":
        desc = f"DIFERENÇA SEGURIDADE SOCIAL - {p.nome} - MATRÍCULA {r.matricula}"
        fill = None
    else:
        desc = f"DEVOLUÇÃO DA SEGURIDADE SOCIAL - {p.nome} - MATRÍCULA {r.matricula}"
        fill = None

    _escrever_celula(ws, linha, 1, p.competencia,
                     _FONTE_DADOS, _DATA_BR, _ALINH_CENTRO,
                     fill=fill, border=_BORDA_DADOS)
    _escrever_celula(ws, linha, 2, p.valor_original,
                     _FONTE_DADOS, _MOEDA, _ALINH_DIR,
                     fill=fill, border=_BORDA_DADOS)
    _escrever_celula(ws, linha, 3, desc,
                     _FONTE_DADOS, alignment=_ALINH_WRAP,
                     fill=fill, border=_BORDA_DADOS)
    _escrever_celula(ws, linha, 4, r.data_alvo,
                     _FONTE_DADOS, _DATA_BR, _ALINH_CENTRO,
                     fill=fill, border=_BORDA_DADOS)
    _escrever_celula(ws, linha, 5, p.correcao,
                     _FONTE_DADOS, _MOEDA, _ALINH_DIR,
                     fill=fill, border=_BORDA_DADOS)
    _escrever_celula(ws, linha, 6, p.fator,
                     _FONTE_DADOS, _FATOR, _ALINH_DIR,
                     fill=fill, border=_BORDA_DADOS)
    _escrever_celula(ws, linha, 7, p.valor_corrigido,
                     _FONTE_DADOS, _MOEDA, _ALINH_DIR,
                     fill=fill, border=_BORDA_DADOS)
